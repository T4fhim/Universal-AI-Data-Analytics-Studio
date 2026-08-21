# File: tests/reports/test_excel_exporter.py
"""Tests for src.reports.excel_exporter.ExcelReportExporter.

Same figure_to_png_bytes-monkeypatching approach as
tests/reports/test_pdf_exporter.py, for the same reason.
"""

from __future__ import annotations

import base64
from pathlib import Path

import plotly.graph_objects as go
import pytest

from src.core.expertise_level import ExpertiseLevel
from src.reports.report_content import ReportContent, ReportSection

_TINY_PNG_BYTES = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY"
    "42YAAAAASUVORK5CYII="
)


@pytest.fixture()
def report_content() -> ReportContent:
    figure = go.Figure(data=[go.Bar(x=["a", "b"], y=[1, 2])])
    return ReportContent(
        title="Sales Report",
        dataset_name="sales",
        dataset_summary={"Rows": 10, "Columns": 3},
        expertise_level=ExpertiseLevel.ANALYST,
        sections=[
            ReportSection(
                stage_label="Analyze: Correlation!?",
                tool_name="compute_correlation",
                timestamp="2026-01-01T00:00:00+00:00",
                summary_lines=["correlation: 0.82"],
                figure=figure,
                explanation_text="Strongly correlated.",
            )
        ],
    )


def test_export_writes_an_xlsx_file(
    tmp_path: Path, report_content: ReportContent, monkeypatch
) -> None:
    monkeypatch.setattr(
        "src.reports.excel_exporter.figure_to_png_bytes",
        lambda *a, **k: _TINY_PNG_BYTES,
    )
    output_path = tmp_path / "report.xlsx"

    from src.reports.excel_exporter import ExcelReportExporter

    result = ExcelReportExporter.export(report_content, output_path)

    assert result == output_path
    assert output_path.exists()
    assert output_path.read_bytes()[:2] == b"PK"  # xlsx is a zip archive


def test_export_sanitizes_sheet_titles_with_invalid_characters(
    tmp_path: Path, report_content: ReportContent, monkeypatch
) -> None:
    monkeypatch.setattr(
        "src.reports.excel_exporter.figure_to_png_bytes",
        lambda *a, **k: _TINY_PNG_BYTES,
    )
    output_path = tmp_path / "report.xlsx"

    from openpyxl import load_workbook

    from src.reports.excel_exporter import ExcelReportExporter

    ExcelReportExporter.export(report_content, output_path)
    workbook = load_workbook(str(output_path))

    assert "Summary" in workbook.sheetnames
    # ":" and "?" and "!" -> "!" is valid, only ":" and "?" get stripped.
    assert any(name.startswith("1_Analyze") for name in workbook.sheetnames)
    assert all(c not in name for name in workbook.sheetnames for c in "[]:*?/\\")
