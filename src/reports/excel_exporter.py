# File: src/reports/excel_exporter.py
"""Static Excel (.xlsx) report exporter via openpyxl.

One worksheet per pipeline stage (plus a Summary sheet), rather than a
single flat sheet: a report's sections vary in count with how many
stages a dataset's log has, and Excel's worksheet-per-topic convention
reads far better for that than jamming everything into one sheet with
visual row separators. Figures are flattened to PNG via
:func:`~src.reports.rasterize.figure_to_png_bytes` and embedded with
openpyxl's own image support — Excel has no notion of an interactive
widget either.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from src.core.exceptions import ServiceError
from src.core.logger import get_logger
from src.reports.base_exporter import BaseReportExporter
from src.reports.rasterize import figure_to_png_bytes
from src.reports.report_content import ReportContent

_logger = get_logger(__name__)

# Excel worksheet titles are capped at 31 characters and cannot contain
# any of []:*?/\ — every character in this set is stripped from a
# stage label before it becomes part of a sheet title.
_INVALID_SHEET_TITLE_CHARS = "[]:*?/\\"


class ExcelReportExporter(BaseReportExporter):
    """Renders a :class:`ReportContent` as a static Excel workbook."""

    @classmethod
    def export(cls, report_content: ReportContent, output_path: Path, **kwargs) -> Path:
        workbook = Workbook()
        cls._render_summary_sheet(workbook, report_content)

        for index, section in enumerate(report_content.sections, start=1):
            cls._render_section_sheet(workbook, section, index)

        try:
            workbook.save(str(output_path))
        except OSError as exc:
            raise ServiceError(
                f"Failed to write Excel report to {output_path}: {exc}"
            ) from exc
        _logger.info("Exported Excel report to %s", output_path)
        return output_path

    @staticmethod
    def _render_summary_sheet(
        workbook: Workbook, report_content: ReportContent
    ) -> None:
        sheet = workbook.active
        sheet.title = "Summary"
        sheet.append(["Report", report_content.title])
        sheet.append(["Dataset", report_content.dataset_name])
        sheet.append(["Generated", report_content.generated_at])
        sheet.append(["Expertise level", report_content.expertise_level.value])
        sheet.append([])
        sheet.append(["Dataset Summary"])
        for key, value in report_content.dataset_summary.items():
            sheet.append([key, value])
        sheet.column_dimensions[get_column_letter(1)].width = 24
        sheet.column_dimensions[get_column_letter(2)].width = 60

    @staticmethod
    def _render_section_sheet(workbook: Workbook, section, index: int) -> None:
        cleaned_label = "".join(
            c for c in section.stage_label if c not in _INVALID_SHEET_TITLE_CHARS
        )
        sheet = workbook.create_sheet(title=f"{index}_{cleaned_label}"[:31])

        sheet.append([section.stage_label])
        if section.tool_name:
            sheet.append(["Tool", section.tool_name])
        sheet.append(["Timestamp", section.timestamp])
        sheet.append([])
        for line in section.summary_lines:
            sheet.append([line])
        if section.explanation_text:
            sheet.append([])
            sheet.append(["Explanation", section.explanation_text])
        sheet.column_dimensions[get_column_letter(1)].width = 60

        if section.figure is not None:
            image_bytes = figure_to_png_bytes(section.figure)
            if image_bytes is not None:
                sheet.add_image(XLImage(BytesIO(image_bytes)), f"A{sheet.max_row + 2}")
